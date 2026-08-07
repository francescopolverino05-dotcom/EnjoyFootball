#!/usr/bin/env python3
"""Convert RPE Excel (CR-10 sheet) → src/data/rpe-load.json."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Install openpyxl: pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "Downloads" / "RPE 2.xlsx"
OUT = ROOT / "src" / "data" / "rpe-load.json"
PLAYERS = ROOT / "src" / "data" / "players.json"

DATE_SLUGS = {
    "2026-07-28": {
        "trainingSlug": "2026-07-28_martedi-forza",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-07-29": {
        "trainingSlug": "2026-07-29_mercoledi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-07-30": {
        "trainingSlug": "2026-07-30_giovedi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-07-31": {
        "trainingSlug": "2026-07-31_venerdi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-08-01": {
        "trainingSlug": None,
        "matchSlug": "2026-08-01_amichevole-u19-vs-u18",
        "kind": "match",
    },
    "2026-08-03": {
        "trainingSlug": "2026-08-03_lunedi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-08-04": {
        "trainingSlug": "2026-08-04_martedi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-08-05": {
        "trainingSlug": "2026-08-05_mercoledi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-08-06": {
        "trainingSlug": "2026-08-06_giovedi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-08-07": {
        "trainingSlug": "2026-08-07_venerdi",
        "matchSlug": None,
        "kind": "training",
    },
}

ALIASES = {
    "ruvetini alessandro": "rovetini-alessandro",
    "saviano raffaele": "saviano-raffaele-junior",
}


def clean_name(s: str) -> str:
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Cf")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s).strip().lower()


def display_name(s: str) -> str:
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Cf")
    return re.sub(r"\s+", " ", s).strip()


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, str) and v.startswith("#"):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def build_roster_map(players: list[dict]) -> dict[str, str]:
    by_full = {clean_name(p["displayName"]): p["slug"] for p in players}
    return by_full


def match_player(raw: str, by_full: dict[str, str]) -> str | None:
    n = clean_name(raw)
    if n in ALIASES:
        return ALIASES[n]
    if n in by_full:
        return by_full[n]
    for full, slug in by_full.items():
        if n == full or n.startswith(full) or full.startswith(n):
            return slug
        nt, ft = set(n.split()), set(full.split())
        if len(nt & ft) >= 2 and n.split()[0] == full.split()[0]:
            return slug
    return None


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--sheet", default="CR-10")
    parser.add_argument("--out", type=Path, default=OUT)
    args = parser.parse_args()

    roster = json.loads(PLAYERS.read_text())["players"]
    by_full = build_roster_map(roster)

    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb[args.sheet]

    sessions_meta = []
    c = 2
    while c <= ws.max_column:
        raw = ws.cell(2, c).value
        if raw is None:
            c += 1
            continue
        text = str(raw).strip()
        m = re.search(r"(\d{2})-(\d{2})-(\d{4})", text)
        if not m:
            c += 1
            continue
        d, mo, y = m.group(1), m.group(2), m.group(3)
        if "PARTITA" in text.upper() and y == "2025":
            y = "2026"
        iso = f"{y}-{mo}-{d}"
        if iso not in DATE_SLUGS:
            print(f"WARNING: unknown date {iso} ({text})")
            c += 3
            continue
        sessions_meta.append(
            {
                "col": c,
                "date": iso,
                "present": ws.cell(3, c).value,
                **DATE_SLUGS[iso],
            }
        )
        c += 3

    unmatched: list[str] = []
    sessions = []

    for sm in sessions_meta:
        col = sm["col"]
        players_out = []
        for r in range(5, 30):
            name = ws.cell(r, 1).value
            if not name or str(name).strip().upper() == "MEDIA":
                continue
            raw_name = str(name)
            rpe_n = num(ws.cell(r, col + 1).value)
            if rpe_n is None:
                continue
            mn_n = num(ws.cell(r, col).value)
            tl_n = num(ws.cell(r, col + 2).value)
            slug = match_player(raw_name, by_full)
            excel = display_name(raw_name)
            if not slug:
                if excel not in unmatched:
                    unmatched.append(excel)
                continue
            session_load = (
                int(round(tl_n))
                if tl_n is not None
                else (
                    int(round(mn_n * rpe_n))
                    if mn_n is not None
                    else None
                )
            )
            players_out.append(
                {
                    "playerSlug": slug,
                    "excelName": excel,
                    "min": int(mn_n) if mn_n is not None else None,
                    "rpe": int(rpe_n) if rpe_n == int(rpe_n) else round(rpe_n, 2),
                    "sessionLoad": session_load,
                }
            )

        avg_min = num(ws.cell(30, col).value)
        avg_rpe = num(ws.cell(30, col + 1).value)
        avg_tl = num(ws.cell(30, col + 2).value)
        answered = [p for p in players_out if p["rpe"] is not None]
        if avg_rpe is None and answered:
            avg_rpe = sum(p["rpe"] for p in answered) / len(answered)
            loads = [p["sessionLoad"] for p in answered if p["sessionLoad"] is not None]
            avg_tl = sum(loads) / len(loads) if loads else None
            mins = [p["min"] for p in answered if p["min"] is not None]
            avg_min = sum(mins) / len(mins) if mins else None

        players_out.sort(
            key=lambda p: (
                -(p["sessionLoad"] or 0),
                -(p["rpe"] or 0),
                p["playerSlug"],
            )
        )
        sessions.append(
            {
                "date": sm["date"],
                "kind": sm["kind"],
                "trainingSlug": sm["trainingSlug"],
                "matchSlug": sm["matchSlug"],
                "playersPresent": (
                    int(sm["present"])
                    if sm["present"] is not None
                    else len(answered)
                ),
                "playersAnswered": len(answered),
                "avgMin": round(avg_min, 1) if avg_min is not None else None,
                "avgRpe": round(avg_rpe, 2) if avg_rpe is not None else None,
                "avgSessionLoad": round(avg_tl, 1) if avg_tl is not None else None,
                "players": players_out,
            }
        )

    # Capture unmatched rows even without RPE (for reporting)
    for r in range(5, 30):
        name = ws.cell(r, 1).value
        if not name or str(name).strip().upper() == "MEDIA":
            continue
        excel = display_name(str(name))
        if not match_player(excel, by_full) and excel not in unmatched:
            unmatched.append(excel)

    out = {
        "source": f"{args.xlsx.name} sheet {args.sheet} (Napoli Primavera ~28 Jul–6 Aug 2026)",
        "scaleNote": {
            "en": "Values look Borg-like (approx. 6–20) despite the CR-10 sheet title. Shown as RPE and session load (min × RPE).",
            "it": "I valori sembrano in scala Borg (circa 6–20) nonostante il titolo foglio CR-10. Mostrati come RPE e carico di seduta (min × RPE).",
        },
        "sessions": sessions,
        "unmatchedNames": unmatched,
    }
    args.out.write_text(json.dumps(out, ensure_ascii=False, indent=2) + "\n")
    print(f"Wrote {args.out} ({len(sessions)} sessions)")
    print("Unmatched:", unmatched or "(none)")


if __name__ == "__main__":
    main()
