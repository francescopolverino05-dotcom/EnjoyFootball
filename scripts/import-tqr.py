#!/usr/bin/env python3
"""Convert TQR Excel → src/data/tqr-load.json.

Copies only TQR values and associated player names from the workbook.
Maps names to roster slugs when there is a clear match; does not invent metrics.
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Install openpyxl: pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "Downloads" / "-TQR SSC NAPOLI.xlsx"
OUT = ROOT / "src" / "data" / "tqr-load.json"
PLAYERS = ROOT / "src" / "data" / "players.json"

DATE_SLUGS = {
    "2026-07-28": {
        "trainingSlug": "2026-07-28_martedi-forza",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-07-29": {
        "trainingSlug": "2026-07-29_mercoledi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-07-30": {
        "trainingSlug": "2026-07-30_giovedi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-07-31": {
        "trainingSlug": "2026-07-31_venerdi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-08-01": {
        "trainingSlug": None,
        "matchSlug": "2026-08-01_amichevole-u19-vs-u18",
        "kind": "match",
    },
    "2026-08-02": {
        # No training/match page on the site for this date yet.
        "trainingSlug": None,
        "matchSlug": None,
        "kind": "training",
    },
    "2026-08-03": {
        "trainingSlug": "2026-08-03_lunedi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-08-04": {
        "trainingSlug": "2026-08-04_martedi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-08-05": {
        "trainingSlug": "2026-08-05_mercoledi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-08-06": {
        "trainingSlug": "2026-08-06_giovedi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-08-07": {
        "trainingSlug": "2026-08-07_venerdi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-08-08": {
        "trainingSlug": None,
        "matchSlug": "2026-08-08_amichevole-u19-vs-portici",
        "kind": "match",
    },
    "2026-08-10": {
        "trainingSlug": "2026-08-10_lunedi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-08-11": {
        "trainingSlug": "2026-08-11_martedi",
        "matchSlug": None,
        "kind": "training",
    },
    "2026-08-12": {
        "trainingSlug": "2026-08-12_mercoledi",
        "matchSlug": None,
        "kind": "training",
    },
}

ALIASES = {
    "ruvetini alessandro": "rovetini-alessandro",
    "saviano raffaele": "saviano-raffaele-junior",
    "chialese": "chiaiese-riccardo",
    "chiaiese": "chiaiese-riccardo",
    "chianese": "chiaiese-riccardo",
}


def clean_name(s: str) -> str:
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Cf")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s).strip().lower()


def display_name(s: str) -> str:
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Cf")
    return re.sub(r"\s+", " ", s).strip()


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, str) and v.startswith("#"):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def iso_date(v) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if v is None:
        return None
    text = str(v).strip()
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", text)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.search(r"(\d{2})[/-](\d{2})[/-](\d{4})", text)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return None


def build_roster_map(players: list[dict]) -> dict[str, str]:
    return {clean_name(p["displayName"]): p["slug"] for p in players}


def match_player(raw: str, by_full: dict[str, str]) -> str | None:
    n = clean_name(raw)
    if n in ALIASES:
        return ALIASES[n]
    if n in by_full:
        return by_full[n]
    for full, slug in by_full.items():
        if n == full or n.startswith(full) or full.startswith(n):
            return slug
        nt, ft = set(n.split()), set(full.split())
        if len(nt & ft) >= 2 and n.split()[0] == full.split()[0]:
            return slug
    return None


def find_tqr_col(ws, date_col: int, max_col: int) -> int | None:
    """Locate the TQR header for a session starting at date_col."""
    # Normal layout: date | RPE | UA | TQR  → TQR at date_col+2
    # Aug-2 layout:  date | RPE | TQR       → TQR at date_col+1
    for offset in (2, 1, 3):
        c = date_col + offset
        if c > max_col:
            continue
        header = ws.cell(5, c).value
        if header is not None and str(header).strip().upper() == "TQR":
            return c
    return None


def discover_sessions(ws) -> list[dict]:
    sessions = []
    unmatched_dates: list[str] = []
    c = 1
    while c <= ws.max_column:
        iso = iso_date(ws.cell(4, c).value)
        if not iso:
            c += 1
            continue
        tqr_col = find_tqr_col(ws, c, ws.max_column)
        if tqr_col is None:
            print(f"WARNING: no TQR column for date {iso} at col {c}")
            c += 1
            continue
        meta = DATE_SLUGS.get(iso)
        if not meta:
            unmatched_dates.append(iso)
            print(f"WARNING: unknown date {iso} — included without site slug")
            meta = {
                "trainingSlug": None,
                "matchSlug": None,
                "kind": "training",
            }
        sessions.append(
            {
                "date": iso,
                "dateCol": c,
                "tqrCol": tqr_col,
                **meta,
            }
        )
        c = tqr_col + 1
    return sessions, unmatched_dates


def parse_sheet(ws, by_full: dict[str, str], unmatched: list[str]) -> tuple[list[dict], list[str]]:
    sessions_meta, unmatched_dates = discover_sessions(ws)
    sessions = []

    for sm in sessions_meta:
        tqr_col = sm["tqrCol"]
        players_out = []
        for r in range(6, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name:
                continue
            if str(name).strip().upper() == "MEDIA":
                continue
            raw_name = str(name)
            tqr_n = num(ws.cell(r, tqr_col).value)
            # Skip empty / zero placeholders — Excel fills 0 when no TQR answered.
            if tqr_n is None or tqr_n == 0:
                continue
            excel = display_name(raw_name)
            slug = match_player(raw_name, by_full)
            if not slug:
                if excel not in unmatched:
                    unmatched.append(excel)
                continue
            players_out.append(
                {
                    "playerSlug": slug,
                    "excelName": excel,
                    "tqr": int(tqr_n) if tqr_n == int(tqr_n) else round(tqr_n, 2),
                }
            )

        media_tqr = num(ws.cell(31, tqr_col).value)
        answered = players_out
        if (media_tqr is None or media_tqr == 0) and answered:
            media_tqr = sum(p["tqr"] for p in answered) / len(answered)

        players_out.sort(key=lambda p: (-(p["tqr"] or 0), p["playerSlug"]))
        sessions.append(
            {
                "date": sm["date"],
                "kind": sm["kind"],
                "trainingSlug": sm["trainingSlug"],
                "matchSlug": sm["matchSlug"],
                "playersAnswered": len(answered),
                "avgTqr": round(media_tqr, 2) if media_tqr is not None else None,
                "players": players_out,
            }
        )

    # Report unmatched names even when all TQR cells were empty/zero.
    for r in range(6, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or str(name).strip().upper() == "MEDIA":
            continue
        excel = display_name(str(name))
        if not match_player(excel, by_full) and excel not in unmatched:
            unmatched.append(excel)

    return sessions, unmatched_dates


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out", type=Path, default=OUT)
    args = parser.parse_args()

    roster = json.loads(PLAYERS.read_text())["players"]
    by_full = build_roster_map(roster)

    wb = load_workbook(args.xlsx, data_only=True)
    unmatched: list[str] = []
    all_sessions: list[dict] = []
    all_unmatched_dates: list[str] = []

    for sheet_name in wb.sheetnames:
        if not str(sheet_name).upper().startswith("TQR"):
            continue
        sessions, unmatched_dates = parse_sheet(wb[sheet_name], by_full, unmatched)
        all_sessions.extend(sessions)
        all_unmatched_dates.extend(unmatched_dates)

    # De-dupe by date (later sheets should not collide; keep first).
    by_date: dict[str, dict] = {}
    for s in all_sessions:
        by_date.setdefault(s["date"], s)
    sessions = sorted(by_date.values(), key=lambda s: s["date"])

    out = {
        "source": f"{args.xlsx.name} sheets TQR 1 + TQR (2) (Napoli Primavera ~28 Jul–6 Aug 2026)",
        "scaleNote": {
            "en": "TQR (Total Quality Recovery) values copied from the Excel as reported (typically ~6–20; higher = better perceived recovery).",
            "it": "Valori TQR (Total Quality Recovery) copiati dall’Excel così come riportati (tipicamente ~6–20; più alto = migliore recupero percepito).",
        },
        "sessions": sessions,
        "unmatchedNames": unmatched,
        "unmatchedDates": sorted(set(all_unmatched_dates)),
    }
    args.out.write_text(json.dumps(out, ensure_ascii=False, indent=2) + "\n")
    print(f"Wrote {args.out} ({len(sessions)} sessions)")
    print("Unmatched names:", unmatched or "(none)")
    print("Unmatched dates:", sorted(set(all_unmatched_dates)) or "(none)")
    for s in sessions:
        print(
            f"  {s['date']}: avgTqr={s['avgTqr']} n={s['playersAnswered']} "
            f"slug={s['trainingSlug'] or s['matchSlug'] or '(no page)'}"
        )


if __name__ == "__main__":
    main()
