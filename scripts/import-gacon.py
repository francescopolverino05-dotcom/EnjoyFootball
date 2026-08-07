#!/usr/bin/env python3
"""Convert Gacon aerobic test Excel → src/data/gacon-load.json.

Copies only VAM finale (km/h) and associated player names from the workbook.
Maps names to roster slugs when there is a clear match; does not invent metrics.
Players without a VAM value are omitted.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import unicodedata
from collections import Counter
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Install openpyxl: pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = (
    Path.home()
    / ".config"
    / "sscn-primavera"
    / "test-di-gacon-napoli.xlsx"
)
OUT = ROOT / "src" / "data" / "gacon-load.json"
PLAYERS = ROOT / "src" / "data" / "players.json"

# Staff date for this Agosto VAM column (Wed 29 Jul 2026).
SESSION_DATE = "2026-07-29"
SESSION_META = {
    "trainingSlug": "2026-07-29_mercoledi",
    "matchSlug": None,
    "kind": "training",
    "period": "agosto",
}

ALIASES = {
    "ruvetini alessandro": "rovetini-alessandro",
    "rovettini": "rovetini-alessandro",
    "rovettini alessandro": "rovetini-alessandro",
    "zappalardo": "zappala-andrea",
    "zappalardo andrea": "zappala-andrea",
}


def clean_name(s: str) -> str:
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Cf")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s).strip().lower()


def display_name(s: str) -> str:
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Cf")
    return re.sub(r"\s+", " ", s).strip()


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, str) and v.startswith("#"):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def build_roster_map(players: list[dict]) -> tuple[dict[str, str], Counter]:
    by_full = {clean_name(p["displayName"]): p["slug"] for p in players}
    surnames = Counter(clean_name(p["displayName"]).split()[0] for p in players)
    return by_full, surnames


def match_player(
    raw: str, by_full: dict[str, str], surnames: Counter
) -> str | None:
    n = clean_name(raw)
    if not n:
        return None
    if n in ALIASES:
        return ALIASES[n]
    if n in by_full:
        return by_full[n]

    # Unique surname-only (e.g. "Rovettini" after alias miss).
    tokens = n.split()
    if len(tokens) == 1 and surnames[tokens[0]] == 1:
        for full, slug in by_full.items():
            if full.split()[0] == tokens[0]:
                return slug

    for full, slug in by_full.items():
        if n == full or n.startswith(full) or full.startswith(n):
            return slug
        nt, ft = set(n.split()), set(full.split())
        if len(nt & ft) >= 2 and n.split()[0] == full.split()[0]:
            return slug
        # Close surname typo with unique target (Zappalardo → Zappalá).
        sur = full.split()[0]
        raw_sur = tokens[0]
        if (
            len(raw_sur) >= 5
            and raw_sur[:5] == sur[:5]
            and abs(len(raw_sur) - len(sur)) <= 3
            and surnames[sur] == 1
            and (len(tokens) == 1 or tokens[0] == raw_sur)
        ):
            return slug
    return None


def parse_sheet(
    ws, by_full: dict[str, str], surnames: Counter
) -> tuple[dict, list[str], list[str]]:
    unmatched: list[str] = []
    players_out: list[dict] = []
    skipped_no_vam: list[str] = []

    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        raw_name = str(name)
        excel = display_name(raw_name)
        if not excel:
            continue

        vam = num(ws.cell(r, 2).value)  # AGOSTO → VAM FINALE (KM/H)
        if vam is None:
            skipped_no_vam.append(excel)
            continue

        slug = match_player(raw_name, by_full, surnames)
        if not slug:
            if excel not in unmatched:
                unmatched.append(excel)
            continue

        players_out.append(
            {
                "playerSlug": slug,
                "excelName": excel,
                "vamKmh": int(vam) if vam == int(vam) else round(vam, 2),
            }
        )

    players_out.sort(key=lambda p: (-(p["vamKmh"] or 0), p["playerSlug"]))
    avg = (
        round(sum(p["vamKmh"] for p in players_out) / len(players_out), 2)
        if players_out
        else None
    )

    session = {
        "date": SESSION_DATE,
        "kind": SESSION_META["kind"],
        "trainingSlug": SESSION_META["trainingSlug"],
        "matchSlug": SESSION_META["matchSlug"],
        "period": SESSION_META["period"],
        "metric": "vamKmh",
        "unit": "km/h",
        "playersTested": len(players_out),
        "avgVamKmh": avg,
        "players": players_out,
    }
    return session, unmatched, skipped_no_vam


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out", type=Path, default=OUT)
    parser.add_argument(
        "--cache-copy",
        type=Path,
        default=None,
        help="Optional path to keep a local copy of the Excel for re-import (outside the repo)",
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"Excel not found: {args.xlsx}")

    if args.cache_copy is not None and args.xlsx.resolve() != args.cache_copy.resolve():
        args.cache_copy.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(args.xlsx, args.cache_copy)
        print(f"Cached Excel → {args.cache_copy}")

    roster = json.loads(PLAYERS.read_text())["players"]
    by_full, surnames = build_roster_map(roster)

    wb = load_workbook(args.xlsx, data_only=True)
    sheet_name = next(
        (n for n in wb.sheetnames if "gacon" in n.lower() or "aerobico" in n.lower()),
        wb.sheetnames[0],
    )
    session, unmatched, skipped = parse_sheet(wb[sheet_name], by_full, surnames)

    out = {
        "source": (
            f"{args.xlsx.name} sheet {sheet_name} — Agosto VAM finale (km/h); "
            f"mapped to {SESSION_DATE} ({SESSION_META['trainingSlug']})"
        ),
        "scaleNote": {
            "en": (
                "Gacon aerobic test — VAM finale (km/h) copied from the staff Excel "
                "(Agosto column). Higher = higher maximal aerobic speed. "
                "Only players with a recorded VAM are listed."
            ),
            "it": (
                "Test aerobico di Gacon — VAM finale (km/h) copiata dall’Excel dello staff "
                "(colonna Agosto). Più alto = velocità aerobica massima più elevata. "
                "Solo i giocatori con VAM registrata sono elencati."
            ),
        },
        "sessions": [session],
        "unmatchedNames": unmatched,
        "skippedNoVam": skipped,
    }
    args.out.write_text(json.dumps(out, ensure_ascii=False, indent=2) + "\n")
    print(f"Wrote {args.out} (1 session, {session['playersTested']} players)")
    print("Unmatched names:", unmatched or "(none)")
    print("Skipped (no VAM):", skipped or "(none)")
    print(
        f"  {session['date']}: avgVam={session['avgVamKmh']} n={session['playersTested']} "
        f"slug={session['trainingSlug']}"
    )
    for p in session["players"]:
        print(f"    {p['playerSlug']}: {p['vamKmh']} km/h ({p['excelName']})")


if __name__ == "__main__":
    main()
